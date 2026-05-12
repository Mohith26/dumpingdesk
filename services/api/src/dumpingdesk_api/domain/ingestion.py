from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Mapping

from dumpingdesk_api.domain.models import CostOfProduction, DatasetKind, HomeMarketSale, USSale


@dataclass(frozen=True)
class FieldMapping:
    columns: Mapping[str, str]

    def source_for(self, canonical_field: str) -> str:
        return self.columns[canonical_field]


def ingest_tabular_file(
    path: Path | str,
    dataset_kind: DatasetKind,
    field_mapping: FieldMapping,
) -> list[USSale] | list[HomeMarketSale] | list[CostOfProduction]:
    rows = _read_rows(Path(path))
    return [_record_from_row(row, dataset_kind, field_mapping) for row in rows]


def _read_rows(path: Path) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(path)
    raise ValueError(f"Unsupported flat-file format: {path.suffix}")


def _read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]


def _record_from_row(
    row: dict[str, object],
    dataset_kind: DatasetKind,
    field_mapping: FieldMapping,
) -> USSale | HomeMarketSale | CostOfProduction:
    if dataset_kind == DatasetKind.COSTS:
        return CostOfProduction(
            connum=_string(row, field_mapping, "connum"),
            variable_cost=_decimal(row, field_mapping, "variable_cost"),
            fixed_cost=_decimal(row, field_mapping, "fixed_cost"),
            currency=_optional_string(row, field_mapping, "currency", "USD"),
        )

    sale_kwargs = {
        "invoice_id": _string(row, field_mapping, "invoice_id"),
        "connum": _string(row, field_mapping, "connum"),
        "gross_unit_price": _decimal(row, field_mapping, "gross_unit_price"),
        "quantity": _decimal(row, field_mapping, "quantity"),
        "sale_date": _date(row, field_mapping, "sale_date"),
        "currency": _optional_string(row, field_mapping, "currency", "USD"),
        "movement_expense": _optional_decimal(row, field_mapping, "movement_expense"),
        "direct_selling_expense": _optional_decimal(row, field_mapping, "direct_selling_expense"),
        "packing_expense": _optional_decimal(row, field_mapping, "packing_expense"),
    }
    if dataset_kind == DatasetKind.US_SALES:
        return USSale(**sale_kwargs)
    if dataset_kind == DatasetKind.HOME_MARKET_SALES:
        return HomeMarketSale(**sale_kwargs)
    raise ValueError(f"Unsupported dataset kind: {dataset_kind}")


def _value(row: dict[str, object], field_mapping: FieldMapping, canonical_field: str) -> object:
    source_column = field_mapping.source_for(canonical_field)
    return row.get(source_column)


def _string(row: dict[str, object], field_mapping: FieldMapping, canonical_field: str) -> str:
    value = _value(row, field_mapping, canonical_field)
    if value is None:
        return ""
    return str(value).strip()


def _optional_string(
    row: dict[str, object],
    field_mapping: FieldMapping,
    canonical_field: str,
    default: str,
) -> str:
    if canonical_field not in field_mapping.columns:
        return default
    value = _string(row, field_mapping, canonical_field)
    return value or default


def _decimal(row: dict[str, object], field_mapping: FieldMapping, canonical_field: str) -> Decimal:
    raw = _value(row, field_mapping, canonical_field)
    try:
        return Decimal(str(raw).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"Invalid decimal for {canonical_field}: {raw}") from exc


def _optional_decimal(
    row: dict[str, object],
    field_mapping: FieldMapping,
    canonical_field: str,
) -> Decimal:
    if canonical_field not in field_mapping.columns:
        return Decimal("0")
    raw = _value(row, field_mapping, canonical_field)
    if raw in (None, ""):
        return Decimal("0")
    return _decimal(row, field_mapping, canonical_field)


def _date(row: dict[str, object], field_mapping: FieldMapping, canonical_field: str) -> date:
    raw = _value(row, field_mapping, canonical_field)
    if isinstance(raw, date):
        return raw
    return date.fromisoformat(str(raw).strip())
